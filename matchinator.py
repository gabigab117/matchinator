import openpyxl
from difflib import get_close_matches

class Matchinator:
    def __init__(self):
        self.rejects = self.rejects_sheet()
        self.orderable = self.orderable_sheet()
        self.rejects_dict = self.create_rejects_dict()
        self.orderable_dict = self.create_orderable_dict()
        self.match_dict = self.create_match_dict()
    
    def rejects_sheet(self):
        wb_rejects = openpyxl.load_workbook("rejets.xlsx")
        rejects_sheet = wb_rejects["rejets"]
        return rejects_sheet
    
    def orderable_sheet(self):
        wb_orderable = openpyxl.load_workbook("cadencier.xlsx")
        orderable_sheet = wb_orderable["Feuil1"]
        return orderable_sheet
    
    def create_rejects_dict(self):
        rejects = {}
        for cel_d, cell_f in zip(self.rejects["D"], self.rejects["F"]):
            rejects[cel_d.value] = cell_f.value
    
        return rejects
    
    def create_orderable_dict(self):
        cadencier = {}
        for cel_l, cell_b in zip(self.orderable["L"], self.orderable["B"]):
            cadencier[cel_l.value] = cell_b.value
        return cadencier
    
    def create_match_dict(self):
        reach = {}
        for v in self.rejects_dict.values():
            matches = get_close_matches(v, self.orderable_dict.values())
            reach[v] = {
                match_item: next(key for key, value in self.orderable_dict.items() if value == match_item)
                for match_item in matches
            }
        return reach
    
    def create_match_workbook(self):
        new_wb = openpyxl.Workbook()
        active_sheet = new_wb.active
        row_index = 1
        for reject_product, dict_interieur in self.match_dict.items():
            active_sheet.cell(row=row_index, column=1, value=reject_product)
            row_index += 1
    
            for (product, code) in dict_interieur.items():
                active_sheet.cell(row=row_index, column=2, value=code)
                active_sheet.cell(row=row_index, column=3, value=product)
                row_index += 1

        new_wb.save("match.xlsx")

